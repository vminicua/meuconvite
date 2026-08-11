from django.contrib import messages
from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Q
from django.http import Http404, HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.templatetags.static import static
from django.contrib.staticfiles import finders
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import segno

from audit.models import AuditAction
from audit.services import log_action, log_create, log_delete, log_update, model_to_dict
from core.ratelimit import rate_limit
from core.utils import get_client_ip
from subscriptions.services import (
    enabled_guest_ids,
    limits,
    select_new_guest_if_capacity,
    set_guest_plan_access,
    sms_count,
)
from weddings.permissions import capability_flags, require_wedding

from .forms import BulkInvitationForm, GiftForm, GuestForm, GuestImportForm, SendInvitationForm
from .models import Gift, GiftSelection, Guest, InvitationChannel, RSVPStatus
from . import messaging


def _guest_invitation_url(request: HttpRequest, guest: Guest) -> str:
    return request.build_absolute_uri(reverse("guest_invitation", args=[guest.invitation_token]))


def _guest_checkin_url(request: HttpRequest, guest: Guest) -> str:
    return request.build_absolute_uri(reverse("guest_checkin", args=[guest.invitation_token]))


def _invitation_rate_key(request: HttpRequest) -> str:
    """Limita cada convite sem bloquear crawlers partilhados entre vários links."""
    return f"{get_client_ip(request) or 'anon'}:{request.path}"


def _share_cover_url(request: HttpRequest, wedding, guest=None) -> str:
    """Imagem simples do desenho para anexar, sem o mockup do catálogo."""
    if guest is not None:
        cover_url = request.build_absolute_uri(
            reverse("guest_invitation_share_image", args=[guest.invitation_token])
        )
        return f"{cover_url}?v={messaging.whatsapp_cover_version(wedding)}"
    if wedding.cover_image:
        return request.build_absolute_uri(wedding.cover_image.url)
    assets = {
        "carta-selada": "img/invitations/burgundy-lace-v2.png",
        "envelope-botanico": "img/invitations/botanical-elegance-v1.webp",
        "classico-dourado": "img/invitations/classic-gold-v1.webp",
        "luxo-preto": "img/invitations/black-gold-v1.webp",
        "capulana": "img/invitations/capulana-editorial-v1.webp",
        "floral-rosa": "img/invitations/floral-terracotta-v2.png",
        "minimal-branco": "img/invitations/minimal-paper-v1.webp",
        "azul-marinho": "img/invitations/navy-silver-v1.webp",
        "terracota": "img/invitations/floral-terracotta-v2.png",
        "tropical": "img/invitations/tropical-editorial-v1.webp",
        "lavanda": "img/invitations/lavender-editorial-v1.webp",
        "areia-dourada": "img/invitations/floral-terracotta-v2.png",
        "noite-estrelada": "img/invitations/starry-night-v1.webp",
    }
    return request.build_absolute_uri(static(assets.get(wedding.selected_template, "img/invitations/classic-gold-v1.webp")))


def _qr_data_uri(url: str) -> str:
    return segno.make(url, error="h").svg_data_uri(scale=5, border=2)


def invitation_default_music(request: HttpRequest) -> HttpResponse:
    """Serve the bundled MP3 with MIME and byte-range support for mobile browsers."""
    audio_path = settings.BASE_DIR / "static" / "audio" / "default-wedding.mp3"
    if not audio_path.exists():
        raise Http404
    content = audio_path.read_bytes()
    total = len(content)
    start, end = 0, total - 1
    status = 200
    range_header = request.headers.get("Range", "")
    if range_header.startswith("bytes="):
        requested = range_header.removeprefix("bytes=").split(",", 1)[0]
        first, _, last = requested.partition("-")
        try:
            start = int(first) if first else 0
            end = int(last) if last else total - 1
        except ValueError:
            start, end = 0, total - 1
        start = max(0, min(start, total - 1))
        end = max(start, min(end, total - 1))
        status = 206
    response = HttpResponse(content[start : end + 1], content_type="audio/mpeg", status=status)
    response["Accept-Ranges"] = "bytes"
    response["Content-Length"] = str(end - start + 1)
    response["Cache-Control"] = "public, max-age=604800"
    if status == 206:
        response["Content-Range"] = f"bytes {start}-{end}/{total}"
    return response


@require_wedding("can_manage_guests")
def guest_list(request: HttpRequest, wedding) -> HttpResponse:
    current_limits = limits(wedding)
    if request.method == "POST":
        form = GuestForm(request.POST, wedding=wedding, allow_seating=current_limits.allows_seating)
        if form.is_valid():
            guest = form.save(commit=False)
            guest.wedding = wedding
            guest.save()
            select_new_guest_if_capacity(wedding=wedding, guest=guest)
            form.instance = guest
            form.save_m2m()
            log_create(guest, actor=request.user, wedding=wedding, request=request)
            messages.success(request, "Convidado acrescentado.")
            return redirect("guests:list", wedding_id=wedding.pk)
    else:
        form = GuestForm(wedding=wedding, allow_seating=current_limits.allows_seating)

    guests = list(
        Guest.objects.filter(wedding=wedding, is_active=True)
        .prefetch_related("allowed_events", "invitation_deliveries")
        .order_by("full_name")
    )
    active_programme_ids = set(
        wedding.events.filter(is_active=True).values_list("pk", flat=True)
    )
    enabled_ids = enabled_guest_ids(wedding, current_limits.max_guests)
    guest_rows = []
    for guest in guests:
        is_enabled = guest.pk in enabled_ids
        invitation_url = _guest_invitation_url(request, guest) if is_enabled else ""
        deliveries = list(guest.invitation_deliveries.all())
        allowed_events = list(guest.allowed_events.all())
        allowed_event_ids = {event.pk for event in allowed_events if event.is_active}
        guest_rows.append({
            "guest": guest,
            "is_enabled": is_enabled,
            "invitation_url": invitation_url,
            "qr_data_uri": _qr_data_uri(_guest_checkin_url(request, guest)) if is_enabled else "",
            "share_message": messaging.invitation_message(
                guest, invitation_url, InvitationChannel.WHATSAPP
            ) if is_enabled else "",
            "share_cover_url": _share_cover_url(request, wedding, guest) if is_enabled else "",
            "allowed_events": allowed_events,
            "has_full_programme": bool(active_programme_ids)
            and allowed_event_ids == active_programme_ids,
            "latest_delivery": deliveries[0] if deliveries else None,
            "edit_form": GuestForm(
                instance=guest,
                wedding=wedding,
                allow_seating=current_limits.allows_seating,
                auto_id=f"edit-{guest.pk}-%s",
            ),
        })
    guest_rows.sort(key=lambda row: (not row["is_enabled"], row["guest"].full_name.casefold()))
    for position, row in enumerate(guest_rows, start=1):
        row["ord"] = position
    return render(
        request,
        "guests/guest_list.html",
        {
            "wedding": wedding,
            "guests": guests,
            "guest_rows": guest_rows,
            "form": form,
            "limits": current_limits,
            "guest_count": len(guests),
            "enabled_count": len(enabled_ids),
            "sms_used": sms_count(wedding),
            "import_form": GuestImportForm(),
            "capabilities": capability_flags(wedding, request.user),
        },
    )


@require_wedding("can_manage_guests")
def guest_export_excel(request: HttpRequest, wedding) -> HttpResponse:
    """Exporta a lista filtrada num ficheiro Excel real (.xlsx)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    queryset = Guest.objects.filter(wedding=wedding, is_active=True).prefetch_related(
        "allowed_events"
    )
    search = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    programme = request.GET.get("programme", "").strip()
    if search:
        queryset = queryset.filter(
            Q(full_name__icontains=search)
            | Q(phone__icontains=search)
            | Q(email__icontains=search)
        )
    if status:
        queryset = queryset.filter(rsvp_status=status)
    if programme:
        queryset = queryset.filter(allowed_events__name__iexact=programme)

    enabled_ids = enabled_guest_ids(wedding)
    guests = list(queryset.distinct())
    guests.sort(key=lambda guest: (guest.pk not in enabled_ids, guest.full_name.casefold()))
    active_ids = set(wedding.events.filter(is_active=True).values_list("pk", flat=True))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Convidados"
    headers = [
        "Ord.", "Convidado", "Telefone", "Email", "Lugares", "Mesa / cadeira",
        "Programa autorizado", "Confirmação", "Estado",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B5903E")

    for position, guest in enumerate(guests, start=1):
        allowed_events = list(guest.allowed_events.all())
        selected_ids = {event.pk for event in allowed_events if event.is_active}
        if active_ids and selected_ids == active_ids:
            programme_label = "Programa completo"
        elif selected_ids:
            programme_label = ", ".join(
                event.name for event in allowed_events if event.pk in selected_ids
            )
        else:
            programme_label = "Sem acesso ao programa"
        sheet.append([
            position,
            guest.full_name,
            guest.phone,
            guest.email,
            guest.party_size,
            guest.seating_assignment,
            programme_label,
            guest.get_rsvp_status_display(),
            "Activo" if guest.pk in enabled_ids else "Bloqueado — requer subscrição",
        ])

    widths = {
        "A": 8, "B": 30, "C": 20, "D": 32, "E": 10, "F": 24,
        "G": 34, "H": 20, "I": 32,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="convidados.xlsx"'
    return response


@require_wedding("can_manage_guests")
def guest_import_template(request: HttpRequest, wedding) -> HttpResponse:
    """Modelo Excel simples e formatado para importação de convidados."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Convidados"
    headers = ["Nome completo*", "Telefone", "Email", "Lugares", "Mesa / cadeira", "Programa", "Notas"]
    sheet.append(headers)
    sheet.append([None, None, None, None, None, None, None])
    sheet["A2"].comment = Comment(
        "Exemplo: Ana Mucavele | +258840000000 | ana@exemplo.com | 2 | Mesa 4 | Todos | Família",
        "MeuConvite",
    )
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B5903E")
    widths = [32, 22, 34, 12, 24, 36, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    seats = DataValidation(type="whole", operator="between", formula1="1", formula2="20")
    seats.error = "Indique um número entre 1 e 20."
    seats.errorTitle = "Lugares inválidos"
    sheet.add_data_validation(seats)
    seats.add("D2:D2001")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:G2001"
    instructions = workbook.create_sheet("Instruções")
    instructions.append(["Como importar convidados"])
    instructions.append(["1. Não altere os nomes das colunas da folha Convidados."])
    instructions.append(["2. Nome completo é obrigatório; telefone ou email são necessários para enviar."])
    instructions.append(["3. Em Programa use Todos, deixe vazio, ou separe momentos por vírgulas."])
    instructions.append(["4. Apague a linha de exemplo antes de carregar o ficheiro."])
    instructions["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    instructions["A1"].fill = PatternFill("solid", fgColor="1F2933")
    instructions.column_dimensions["A"].width = 95
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="modelo-convidados.xlsx"'
    return response


@require_POST
@require_wedding("can_manage_guests")
def guest_import_excel(request: HttpRequest, wedding) -> HttpResponse:
    """Valida todo o Excel antes de criar convidados, evitando importações parciais."""
    from openpyxl import load_workbook

    form = GuestImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, " ".join(error for errors in form.errors.values() for error in errors))
        return redirect("guests:list", wedding_id=wedding.pk)
    try:
        workbook = load_workbook(form.cleaned_data["file"], read_only=True, data_only=True)
        sheet = workbook["Convidados"] if "Convidados" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
    except Exception as exc:
        messages.error(request, "Não foi possível ler este ficheiro Excel.")
        return redirect("guests:list", wedding_id=wedding.pk)
    if not raw_headers:
        messages.error(request, "O ficheiro está vazio.")
        return redirect("guests:list", wedding_id=wedding.pk)

    def header_key(value):
        import unicodedata
        text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
        return " ".join(text.replace("*", "").replace("_", " ").split())

    aliases = {
        "nome completo": "full_name", "nome": "full_name", "convidado": "full_name",
        "telefone": "phone", "telemovel": "phone", "email": "email",
        "lugares": "party_size", "mesa / cadeira": "seating_assignment",
        "mesa": "seating_assignment", "programa": "programme", "notas": "notes",
    }
    columns = {index: aliases.get(header_key(value)) for index, value in enumerate(raw_headers)}
    if "full_name" not in columns.values():
        messages.error(request, "Falta a coluna obrigatória “Nome completo”. Use o modelo MeuConvite.")
        return redirect("guests:list", wedding_id=wedding.pk)

    active_events = list(wedding.events.filter(is_active=True))
    events_by_name = {event.name.casefold(): event for event in active_events}
    existing_emails = set(
        Guest.objects.filter(wedding=wedding, is_active=True).exclude(email="")
        .values_list("email", flat=True)
    )
    existing_phones = set(
        Guest.objects.filter(wedding=wedding, is_active=True).exclude(phone="")
        .values_list("phone", flat=True)
    )
    parsed, errors, seen_contacts = [], [], set()
    for row_number, values in enumerate(rows, start=2):
        data = {field: values[index] for index, field in columns.items() if field and index < len(values)}
        if not any(value not in (None, "") for value in data.values()):
            continue
        if row_number > 2001:
            errors.append("O ficheiro excede o limite de 2 000 convidados.")
            break
        name = str(data.get("full_name") or "").strip()
        phone_value = data.get("phone")
        phone = str(int(phone_value) if isinstance(phone_value, float) and phone_value.is_integer() else phone_value or "").strip()
        email = str(data.get("email") or "").strip().lower()
        try:
            party_size = int(data.get("party_size") or 1)
        except (TypeError, ValueError):
            party_size = 0
        if not name:
            errors.append(f"Linha {row_number}: o nome completo é obrigatório.")
        if email:
            try: validate_email(email)
            except ValidationError: errors.append(f"Linha {row_number}: email inválido ({email}).")
            if email in existing_emails:
                errors.append(f"Linha {row_number}: já existe um convidado com o email {email}.")
        if phone and phone in existing_phones:
            errors.append(f"Linha {row_number}: já existe um convidado com o telefone {phone}.")
        if not 1 <= party_size <= 20:
            errors.append(f"Linha {row_number}: lugares deve estar entre 1 e 20.")
        contact_key = (email, phone)
        if (email or phone) and contact_key in seen_contacts:
            errors.append(f"Linha {row_number}: contacto repetido no ficheiro.")
        seen_contacts.add(contact_key)
        programme_text = str(data.get("programme") or "Todos").strip()
        if not programme_text or programme_text.casefold() in {"todos", "programa completo"}:
            selected_events = active_events
        else:
            names = [part.strip().casefold() for part in programme_text.split(",") if part.strip()]
            unknown = [part for part in names if part not in events_by_name]
            if unknown:
                errors.append(f"Linha {row_number}: momentos desconhecidos: {', '.join(unknown)}.")
            selected_events = [events_by_name[part] for part in names if part in events_by_name]
        parsed.append({
            "full_name": name, "phone": phone, "email": email, "party_size": party_size,
            "seating_assignment": str(data.get("seating_assignment") or "").strip(),
            "notes": str(data.get("notes") or "").strip(), "events": selected_events,
        })
    if not parsed and not errors:
        errors.append("O ficheiro não contém convidados para importar.")
    if errors:
        preview = " ".join(errors[:8])
        if len(errors) > 8: preview += f" E mais {len(errors) - 8} erro(s)."
        messages.error(request, preview)
        return redirect("guests:list", wedding_id=wedding.pk)

    created = 0
    with transaction.atomic():
        for data in parsed:
            selected_events = data.pop("events")
            guest = Guest.objects.create(wedding=wedding, **data)
            select_new_guest_if_capacity(wedding=wedding, guest=guest)
            guest.allowed_events.set(selected_events)
            created += 1
        log_action(
            action=AuditAction.CREATE, actor=request.user, wedding=wedding, request=request,
            instance=wedding, new_data={"bulk_guest_import": created},
        )
    messages.success(request, f"{created} convidado(s) importado(s) do Excel.")
    return redirect("guests:list", wedding_id=wedding.pk)


@require_POST
@require_wedding("can_manage_guests")
def guest_bulk_send(request: HttpRequest, wedding) -> HttpResponse:
    form = BulkInvitationForm(request.POST, wedding=wedding)
    if not form.is_valid():
        messages.error(request, "Seleccione convidados e um canal de envio.")
        return redirect("guests:list", wedding_id=wedding.pk)
    selected_ids = form.cleaned_data["guest_ids"][:500]
    enabled = enabled_guest_ids(wedding)
    guests = list(Guest.objects.filter(pk__in=selected_ids, wedding=wedding, is_active=True).order_by("full_name"))
    guests = [guest for guest in guests if guest.pk in enabled]
    channel = form.cleaned_data["channel"]
    if channel == InvitationChannel.WHATSAPP:
        rows = []
        for guest in guests:
            try:
                url = messaging.whatsapp_invitation_url(guest=guest, invitation_url=_guest_invitation_url(request, guest))
                rows.append({"guest": guest, "url": url})
            except ValidationError:
                continue
        return render(request, "guests/bulk_whatsapp.html", {
            "wedding": wedding, "rows": rows, "capabilities": capability_flags(wedding, request.user),
        })

    sent, failed = 0, []
    for guest in guests:
        try:
            delivery = messaging.send_invitation(
                guest=guest, channel=channel,
                invitation_url=_guest_invitation_url(request, guest), actor=request.user,
            )
        except ValidationError as exc:
            failed.append(f"{guest.full_name}: {' '.join(exc.messages)}")
            continue
        log_action(
            action=AuditAction.INVITE_SENT, actor=request.user, wedding=wedding,
            request=request, instance=delivery,
            new_data={"channel": delivery.channel, "status": delivery.status, "bulk": True},
        )
        sent += 1
    if sent:
        messages.success(request, f"{sent} convite(s) enviado(s) por {channel}.")
    if failed:
        messages.warning(request, f"{len(failed)} envio(s) falharam. " + " ".join(failed[:3]))
    return redirect("guests:list", wedding_id=wedding.pk)


@require_wedding("can_manage_guests")
def guest_edit(request: HttpRequest, wedding, guest_id) -> HttpResponse:
    guest = get_object_or_404(Guest, pk=guest_id, wedding=wedding, is_active=True)
    allow_seating = limits(wedding).allows_seating
    if request.method == "POST":
        old_data = model_to_dict(guest)
        form = GuestForm(request.POST, instance=guest, wedding=wedding, allow_seating=allow_seating)
        if form.is_valid():
            form.save()
            log_update(guest, old_data, actor=request.user, wedding=wedding, request=request)
            messages.success(request, "Convidado actualizado.")
            return redirect("guests:list", wedding_id=wedding.pk)
    else:
        form = GuestForm(instance=guest, wedding=wedding, allow_seating=allow_seating)
    return render(
        request,
        "guests/guest_form.html",
        {"wedding": wedding, "guest": guest, "form": form,
         "capabilities": capability_flags(wedding, request.user)},
    )


@require_POST
@require_wedding("can_manage_guests")
def guest_send_invitation(request: HttpRequest, wedding, guest_id) -> HttpResponse:
    guest = get_object_or_404(Guest, pk=guest_id, wedding=wedding, is_active=True)
    if guest.pk not in enabled_guest_ids(wedding):
        messages.error(request, "Este convidado requer uma subscrição activa antes do envio.")
        return redirect("guests:list", wedding_id=wedding.pk)

    form = SendInvitationForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Escolha SMS, WhatsApp ou email para enviar o convite.")
        return redirect("guests:list", wedding_id=wedding.pk)

    invitation_url = _guest_invitation_url(request, guest)
    if form.cleaned_data["channel"] == "whatsapp":
        try:
            return redirect(
                messaging.whatsapp_invitation_url(
                    guest=guest,
                    invitation_url=invitation_url,
                )
            )
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
            return redirect("guests:list", wedding_id=wedding.pk)

    try:
        delivery = messaging.send_invitation(
            guest=guest,
            channel=form.cleaned_data["channel"],
            invitation_url=invitation_url,
            actor=request.user,
        )
    except ValidationError as exc:
        messages.error(request, " ".join(exc.messages))
    else:
        log_action(
            action=AuditAction.INVITE_SENT,
            actor=request.user,
            wedding=wedding,
            request=request,
            instance=delivery,
            new_data={"channel": delivery.channel, "status": delivery.status},
        )
        messages.success(
            request,
            f"Convite de {guest.full_name} colocado na fila por {delivery.get_channel_display()}.",
        )
    return redirect("guests:list", wedding_id=wedding.pk)


@require_POST
@require_wedding("can_manage_guests")
def guest_remove(request: HttpRequest, wedding, guest_id) -> HttpResponse:
    guest = get_object_or_404(Guest, pk=guest_id, wedding=wedding, is_active=True)
    log_delete(guest, actor=request.user, wedding=wedding, request=request)
    guest.is_active = False
    guest.save(update_fields=["is_active", "updated_at"])
    messages.info(request, "Convidado removido.")
    return redirect("guests:list", wedding_id=wedding.pk)


@require_POST
@require_wedding("can_manage_guests")
def guest_plan_access(request: HttpRequest, wedding, guest_id) -> HttpResponse:
    guest = get_object_or_404(Guest, pk=guest_id, wedding=wedding, is_active=True)
    action = request.POST.get("action")
    if action not in {"enable", "disable"}:
        messages.error(request, "Escolha incluir ou retirar o convidado do plano.")
        return redirect("guests:list", wedding_id=wedding.pk)
    enabled = action == "enable"
    old_data = model_to_dict(guest)
    try:
        set_guest_plan_access(wedding=wedding, guest=guest, enabled=enabled)
    except ValidationError as exc:
        messages.error(request, " ".join(exc.messages))
    else:
        log_update(guest, old_data, actor=request.user, wedding=wedding, request=request)
        if enabled:
            messages.success(request, f"{guest.full_name} incluído no limite do plano.")
        else:
            messages.info(request, f"{guest.full_name} retirado do limite do plano.")
    return redirect("guests:list", wedding_id=wedding.pk)


@require_wedding("can_manage_guests")
def gift_list(request: HttpRequest, wedding) -> HttpResponse:
    if request.method == "POST":
        form = GiftForm(request.POST)
        if form.is_valid():
            gift = form.save(commit=False)
            gift.wedding = wedding
            gift.save()
            log_create(gift, actor=request.user, wedding=wedding, request=request)
            messages.success(request, "Presente adicionado à lista.")
            return redirect("guests:gifts", wedding_id=wedding.pk)
        messages.error(request, "Corrija os campos assinalados.")
    else:
        form = GiftForm()

    gifts = Gift.objects.filter(wedding=wedding, is_active=True).prefetch_related(
        "selections__guest"
    )
    return render(request, "guests/gift_list.html", {
        "wedding": wedding,
        "form": form,
        "gifts": gifts,
        "capabilities": capability_flags(wedding, request.user),
    })


@require_wedding("can_manage_guests")
def gift_edit(request: HttpRequest, wedding, gift_id) -> HttpResponse:
    gift = get_object_or_404(Gift, pk=gift_id, wedding=wedding, is_active=True)
    if request.method == "POST":
        form = GiftForm(request.POST, instance=gift)
        if form.is_valid():
            old_data = model_to_dict(gift)
            form.save()
            log_update(gift, old_data=old_data, actor=request.user, wedding=wedding, request=request)
            messages.success(request, "Presente actualizado.")
            return redirect("guests:gifts", wedding_id=wedding.pk)
        messages.error(request, "Corrija os campos assinalados.")
    else:
        form = GiftForm(instance=gift)
    return render(request, "guests/gift_form.html", {
        "wedding": wedding, "gift": gift, "form": form,
        "capabilities": capability_flags(wedding, request.user),
    })


@require_POST
@require_wedding("can_manage_guests")
def gift_remove(request: HttpRequest, wedding, gift_id) -> HttpResponse:
    gift = get_object_or_404(Gift, pk=gift_id, wedding=wedding, is_active=True)
    log_delete(gift, actor=request.user, wedding=wedding, request=request)
    gift.is_active = False
    gift.save(update_fields=["is_active", "updated_at"])
    messages.info(request, "Presente removido da lista.")
    return redirect("guests:gifts", wedding_id=wedding.pk)


@rate_limit("invitation_view", methods=("GET",), key_func=_invitation_rate_key)
@rate_limit("rsvp_submit", methods=("POST",))
def guest_invitation(request: HttpRequest, token: str) -> HttpResponse:
    """Individual invitation and RSVP surface addressed by an opaque token."""
    from templates_manager import registry
    from templates_manager.services import invitation_context

    guest = get_object_or_404(
        Guest.objects.select_related("wedding", "wedding__category").prefetch_related("allowed_events"),
        invitation_token=token,
        is_active=True,
    )
    wedding = guest.wedding
    if wedding.status in {"archived", "blocked"}:
        raise Http404
    if guest.pk not in enabled_guest_ids(wedding):
        raise Http404

    if request.method == "POST":
        response = request.POST.get("rsvp")
        if response in {RSVPStatus.CONFIRMED, RSVPStatus.DECLINED}:
            guest.rsvp_status = response
            guest.responded_at = timezone.now()
            guest.save(update_fields=["rsvp_status", "responded_at", "updated_at"])
            messages.success(request, "A sua resposta foi registada. Obrigado!")
            return redirect("guest_invitation", token=guest.invitation_token)

    template = registry.get_template(wedding.selected_template)
    if template is None:
        raise Http404
    context = invitation_context(wedding, template, guest=guest, is_preview=False)
    gifts = list(Gift.objects.filter(wedding=wedding, is_active=True).prefetch_related(
        "selections__guest"
    ))
    for gift in gifts:
        selections = list(gift.selections.all())
        gift.selected_by_guest = any(item.guest_id == guest.pk for item in selections)
        gift.unavailable = bool(selections) and not gift.allow_multiple and not gift.selected_by_guest
    context["gifts"] = gifts
    context["share_cover_url"] = _share_cover_url(request, wedding, guest)
    context["invitation_url"] = _guest_invitation_url(request, guest)
    context["css_variables"] = template.css_variables(
        wedding.primary_color, wedding.secondary_color
    )
    return render(request, "invitations/preview.html", context)


@rate_limit("invitation_view", methods=("GET",), key_func=_invitation_rate_key)
def guest_checkin(request: HttpRequest, token: str) -> HttpResponse:
    """Credencial de entrada aberta exclusivamente pelo QR individual."""
    guest = get_object_or_404(
        Guest.objects.select_related("wedding", "wedding__category").prefetch_related(
            "allowed_events"
        ),
        invitation_token=token,
        is_active=True,
    )
    wedding = guest.wedding
    if wedding.status in {"archived", "blocked"} or guest.pk not in enabled_guest_ids(wedding):
        raise Http404
    qr_events = list(
        guest.allowed_events.filter(is_active=True, requires_qr_code=True)
        .select_related("location")
        .order_by("date", "start_time", "display_order")
    )
    return render(request, "guests/checkin_card.html", {
        "guest": guest,
        "wedding": wedding,
        "qr_events": qr_events,
        "is_demo": False,
    })


@rate_limit("invitation_view", methods=("GET",), key_func=_invitation_rate_key)
def guest_checkin_demo(request: HttpRequest, token: str) -> HttpResponse:
    from weddings.models import Wedding

    wedding = get_object_or_404(
        Wedding.objects.select_related("category"), public_token=token
    )
    if wedding.status in {"archived", "blocked"}:
        raise Http404
    return render(request, "guests/checkin_card.html", {
        "guest": None,
        "wedding": wedding,
        "qr_events": list(
            wedding.events.filter(is_active=True, requires_qr_code=True)
            .select_related("location").order_by("date", "start_time", "display_order")
        ),
        "is_demo": True,
    })


def guest_invitation_share_image(request: HttpRequest, token: str) -> HttpResponse:
    """Cartão JPEG Open Graph personalizado para WhatsApp e outras redes.

    Este recurso não usa o limite por IP da página do convite: os crawlers
    sociais partilham poucos endereços IP e precisam de pedir a página e a
    imagem em sequência para construir a pré-visualização.
    """
    from io import BytesIO

    from PIL import Image, ImageDraw, ImageFilter, ImageFont, ImageOps

    def load_font(size: int, *, serif: bool = False, bold: bool = False):
        """Usa fontes presentes no cPanel e mantém fallback portátil nos testes."""
        if serif:
            names = (
                "/usr/share/fonts/urw-base35/NimbusRoman-Bold.otf" if bold
                else "/usr/share/fonts/urw-base35/NimbusRoman-Regular.otf",
                "C:/Windows/Fonts/georgiab.ttf" if bold else "C:/Windows/Fonts/georgia.ttf",
            )
        else:
            names = (
                "/usr/share/fonts/google-droid/DroidSans-Bold.ttf" if bold
                else "/usr/share/fonts/google-droid/DroidSans.ttf",
                "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
            )
        for name in names:
            try:
                return ImageFont.truetype(name, size=size)
            except OSError:
                continue
        return ImageFont.load_default(size=size)

    def fitted_font(draw, text: str, max_width: int, start: int, minimum: int, **kwargs):
        for size in range(start, minimum - 1, -2):
            font = load_font(size, **kwargs)
            if draw.textbbox((0, 0), text, font=font)[2] <= max_width:
                return font
        return load_font(minimum, **kwargs)

    guest = get_object_or_404(
        Guest.objects.select_related("wedding"), invitation_token=token, is_active=True
    )
    wedding = guest.wedding
    if wedding.status in {"archived", "blocked"} or guest.pk not in enabled_guest_ids(wedding):
        raise Http404

    source = None
    if wedding.cover_image:
        try:
            source = wedding.cover_image.open("rb")
        except (FileNotFoundError, OSError):
            source = None
    if source is None:
        assets = {
            "carta-selada": "img/invitations/burgundy-lace-v2.png",
            "envelope-botanico": "img/invitations/botanical-elegance-v1.webp",
            "classico-dourado": "img/invitations/classic-gold-v1.webp",
            "luxo-preto": "img/invitations/black-gold-v1.webp",
            "capulana": "img/invitations/capulana-editorial-v1.webp",
            "floral-rosa": "img/invitations/floral-terracotta-v2.png",
            "minimal-branco": "img/invitations/minimal-paper-v1.webp",
            "azul-marinho": "img/invitations/navy-silver-v1.webp",
            "terracota": "img/invitations/floral-terracotta-v2.png",
            "tropical": "img/invitations/tropical-editorial-v1.webp",
            "lavanda": "img/invitations/lavender-editorial-v1.webp",
            "areia-dourada": "img/invitations/floral-terracotta-v2.png",
            "noite-estrelada": "img/invitations/starry-night-v1.webp",
        }
        source = finders.find(
            assets.get(wedding.selected_template, "img/invitations/classic-gold-v1.webp")
        )
    if not source:
        raise Http404

    try:
        with Image.open(source) as image:
            image = ImageOps.exif_transpose(image).convert("RGB")
            image = ImageOps.fit(image, (1200, 630), method=Image.Resampling.LANCZOS)
            image = image.filter(ImageFilter.GaussianBlur(2.2)).convert("RGBA")

            # Fundo de veludo escuro e envelope marfim inspirado em
            # correspondência real. As formas são desenhadas no servidor para
            # que cada convidado receba uma capa verdadeiramente personalizada.
            image.alpha_composite(Image.new("RGBA", image.size, (20, 8, 17, 174)))
            draw = ImageDraw.Draw(image, "RGBA")
            draw.rounded_rectangle((58, 58, 1148, 574), radius=28, fill=(0, 0, 0, 105))
            envelope = (70, 46, 1130, 556)
            draw.rounded_rectangle(
                envelope, radius=24, fill=(243, 232, 202, 255),
                outline=(186, 139, 48, 255), width=5,
            )
            draw.rounded_rectangle(
                (84, 60, 1116, 542), radius=18,
                outline=(164, 117, 35, 190), width=2,
            )

            burgundy = (78, 18, 25, 255)
            deep_burgundy = (53, 10, 18, 255)
            gold = (166, 116, 28, 255)
            pale_gold = (206, 165, 80, 255)

            # Cantos ornamentais com linhas e pequenos medalhões dourados.
            for sx, sy in ((1, 1), (-1, 1), (1, -1), (-1, -1)):
                x = 110 if sx == 1 else 1090
                y = 84 if sy == 1 else 518
                draw.line((x, y, x + sx * 82, y), fill=gold, width=3)
                draw.line((x, y, x, y + sy * 54), fill=gold, width=3)
                draw.arc(
                    (x - 12, y - 12, x + 12, y + 12), 0, 360,
                    fill=pale_gold, width=3,
                )
                draw.ellipse(
                    (x + sx * 88 - 4, y - 4, x + sx * 88 + 4, y + 4),
                    fill=pale_gold,
                )

            def centered_text(y, text, font, fill):
                box = draw.textbbox((0, 0), text, font=font)
                x = 600 - (box[2] - box[0]) / 2
                draw.text((x, y), text, font=font, fill=fill)

            kicker_font = load_font(19, bold=True)
            centered_text(88, "CORRESPONDÊNCIA REAL  •  CONVITE PARTICULAR", kicker_font, gold)

            guest_text = (guest.full_name or "Convidado").upper()
            guest_font = fitted_font(draw, guest_text, 720, 31, 22, bold=True)
            centered_text(130, f"PARA  {guest_text}", guest_font, burgundy)

            draw.line((310, 184, 890, 184), fill=(166, 116, 28, 175), width=2)
            draw.ellipse((294, 179, 304, 189), fill=gold)
            draw.ellipse((896, 179, 906, 189), fill=gold)

            names = wedding.display_names
            names_font = fitted_font(draw, names, 850, 66, 42, serif=True, bold=True)
            centered_text(194, names, names_font, deep_burgundy)

            date_text = wedding.main_date.strftime("%d.%m.%Y") if wedding.main_date else ""
            detail_parts = [str(wedding.category_name), date_text, wedding.city]
            detail_text = "  •  ".join(part for part in detail_parts if part)
            detail_font = fitted_font(draw, detail_text.upper(), 720, 25, 18, bold=True)
            centered_text(278, detail_text.upper(), detail_font, gold)

            # Dobras do envelope terminam no selo central sem competir com o texto.
            fold_y = 374
            draw.polygon(
                ((88, 540), (88, 330), (600, fold_y)),
                fill=(226, 210, 174, 225), outline=(177, 135, 58, 170),
            )
            draw.polygon(
                ((1112, 540), (1112, 330), (600, fold_y)),
                fill=(234, 219, 184, 225), outline=(177, 135, 58, 170),
            )
            draw.polygon(
                ((88, 540), (1112, 540), (600, fold_y)),
                fill=(239, 225, 191, 250), outline=(177, 135, 58, 190),
            )

            seal_box = (505, 338, 695, 528)
            seal_path = finders.find("img/invitations/burgundy-wax-seal-v1.png")
            if seal_path:
                with Image.open(seal_path) as seal:
                    seal = seal.convert("RGBA").resize((190, 190), Image.Resampling.LANCZOS)
                    image.alpha_composite(seal, (seal_box[0], seal_box[1]))
            else:
                draw.ellipse(seal_box, fill=(111, 24, 18, 245), outline=gold, width=4)

            initials = "".join(
                name[:1].upper()
                for name in (wedding.primary_short_name, wedding.secondary_short_name)
                if name
            ) or "MC"
            monogram_font = fitted_font(draw, initials, 105, 54, 36, serif=True, bold=True)
            monogram_box = draw.textbbox((0, 0), initials, font=monogram_font)
            monogram_x = 600 - (monogram_box[2] - monogram_box[0]) / 2
            monogram_y = 433 - (monogram_box[3] - monogram_box[1]) / 2 - monogram_box[1]
            draw.text((monogram_x, monogram_y), initials, font=monogram_font, fill=(247, 213, 169, 255))

            callout_font = load_font(15, bold=True)
            centered_text(
                515, "ABRA O CONVITE E CONFIRME A SUA PRESENÇA",
                callout_font, burgundy,
            )
            brand_font = load_font(15, bold=True)
            centered_text(588, "MEUCONVITE.CO.MZ", brand_font, pale_gold)

            output = BytesIO()
            image.convert("RGB").save(
                output, format="JPEG", quality=90, optimize=True, progressive=True
            )
    finally:
        if hasattr(source, "close"):
            source.close()

    response = HttpResponse(output.getvalue(), content_type="image/jpeg")
    response["Cache-Control"] = "public, max-age=86400"
    response["Content-Disposition"] = 'inline; filename="capa-convite.jpg"'
    response["X-Share-Image-Version"] = messaging.WHATSAPP_PREVIEW_REVISION
    return response


@require_POST
@rate_limit("rsvp_submit", methods=("POST",))
def guest_gift_select(request: HttpRequest, token: str, gift_id) -> HttpResponse:
    guest = get_object_or_404(Guest, invitation_token=token, is_active=True)
    wedding = guest.wedding
    if wedding.status in {"archived", "blocked"} or guest.pk not in enabled_guest_ids(wedding):
        raise Http404

    with transaction.atomic():
        gift = get_object_or_404(
            Gift.objects.select_for_update(), pk=gift_id, wedding=wedding, is_active=True
        )
        own_selection = GiftSelection.objects.filter(gift=gift, guest=guest).first()
        if own_selection:
            own_selection.delete()
            messages.info(request, f"Deixou de levar “{gift.name}”.")
        elif not gift.allow_multiple and GiftSelection.objects.filter(gift=gift).exists():
            messages.error(request, "Este presente já foi escolhido por outro convidado.")
        else:
            GiftSelection.objects.create(gift=gift, guest=guest)
            messages.success(request, f"Obrigado! Ficou registado que vai levar “{gift.name}”.")

    return redirect("guest_invitation", token=guest.invitation_token)


@csrf_exempt
@require_POST
def twilio_message_status(request: HttpRequest) -> HttpResponse:
    """Webhook assinado do Twilio para estados de entrega e leitura."""
    from platform_admin.models import configured_value

    auth_token = configured_value("twilio_auth_token")
    signature = request.headers.get("X-Twilio-Signature", "")
    if not auth_token or not signature:
        return HttpResponse(status=403)

    from twilio.request_validator import RequestValidator

    callback_url = f"{settings.SITE_BASE_URL.rstrip('/')}{request.path}"
    if not RequestValidator(auth_token).validate(callback_url, request.POST, signature):
        return HttpResponse(status=403)
    messaging.update_delivery_status(
        provider_sid=request.POST.get("MessageSid", ""),
        status=request.POST.get("MessageStatus", ""),
        error_code=request.POST.get("ErrorCode", ""),
    )
    return HttpResponse(status=204)
